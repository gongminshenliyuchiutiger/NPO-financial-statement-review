import openpyxl
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher
import re

class ReportStandardizer:
    def __init__(self, target_template_path="Full_Financial_Report_Template.xlsx"):
        self.template_path = target_template_path
        
    def normalize_string(self, s):
        if not s: return ""
        return str(s).strip().replace(" ", "").replace("\u3000", "")

    def fuzzy_match(self, source, target, threshold=0.6):
        source = self.normalize_string(source)
        target = self.normalize_string(target)
        if not source or not target: return False
        return SequenceMatcher(None, source, target).ratio() > threshold

    def standardize(self, input_wb):
        # Create a new workbook based on the template structure (conceptually)
        # For simplicity, we create a new WB and manually populate the known structure
        output_wb = openpyxl.Workbook()
        
        # We need to find data for:
        # 1. Income Statement: [Item, LastYear, Budget, ThisYear, Note]
        # 2. Balance Sheet: [Item, LastYear, ThisYear, Note]
        
        self._process_income_statement(input_wb, output_wb)
        self._process_balance_sheet(input_wb, output_wb)
        self._process_property_catalog(input_wb, output_wb)
        
        return output_wb

    def _find_sheet_by_keyword(self, wb, keywords):
        for sheet_name in wb.sheetnames:
            norm_name = self.normalize_string(sheet_name)
            if any(k in norm_name for k in keywords):
                return wb[sheet_name]
        return None

    def _extract_data(self, ws, required_cols):
        # Intelligent column mapping
        # required_cols = {"this_year": ["本年", "決算"], "last_year": ["上年", "去年"], "budget": ["預算"], "item": ["科目", "摘要"]}
        
        header_row_idx = -1
        col_mapping = {} # key (e.g. 'this_year') -> col_index (1-based)

        # 1. Find Header Row (scan first 10 rows)
        for r in range(1, min(11, ws.max_row + 1)):
            row_values = [str(cell.value or "") for cell in ws[r]]
            matches = 0
            temp_mapping = {}
            
            for key, keywords in required_cols.items():
                for col_idx, cell_val in enumerate(row_values, 1):
                    norm_val = self.normalize_string(str(cell_val))
                    if any(kw in norm_val for kw in keywords):
                        temp_mapping[key] = col_idx
                        break
            
            if len(temp_mapping) >= 2: # At least find Item and one amount
                header_row_idx = r
                col_mapping = temp_mapping
                break
        
        if header_row_idx == -1:
            print(f"Warning: Could not identify headers in sheet {ws.title}")
            return []

        # 2. Extract Data
        extracted_data = [] # List of dicts {key: value}
        for r in range(header_row_idx + 1, ws.max_row + 1):
            row_data = {}
            has_data = False
            
            # Get Item Name
            if "item" in col_mapping:
                item_name = ws.cell(row=r, column=col_mapping["item"]).value
                if item_name:
                    row_data["item"] = item_name
                    has_data = True
            
            if not has_data: continue

            # Get Amounts
            for key, col_idx in col_mapping.items():
                if key == "item": continue
                val = ws.cell(row=r, column=col_idx).value
                # Clean number
                try:
                    if isinstance(val, (int, float)):
                        row_data[key] = val
                    elif val and isinstance(val, str):
                        # Remove commas, handle parenthesis as negative
                        clean_val = val.replace(",", "").strip()
                        if clean_val.startswith("(") and clean_val.endswith(")"):
                            clean_val = "-" + clean_val[1:-1]
                        row_data[key] = float(clean_val)
                    else:
                        row_data[key] = 0
                except:
                    row_data[key] = 0
            
            extracted_data.append(row_data)
        
        return extracted_data

    def _process_income_statement(self, input_wb, output_wb):
        ws_in = self._find_sheet_by_keyword(input_wb, ["收支", "損益", "決算"])
        ws_out = output_wb.create_sheet("收支決算表")
        ws_out.append(["科目名稱", "上年度決算數", "本年度預算數", "本年度決算數", "說明"]) # Standard Header
        
        if not ws_in:
            return

        data = self._extract_data(ws_in, {
            "item": ["科目", "項目", "摘要", "名稱"],
            "this_year": ["本年", "本期", "決算"],
            "last_year": ["上年", "上期", "去年", "比較"],
            "budget": ["預算"]
        })

        for row in data:
            ws_out.append([
                row.get("item", ""),
                row.get("last_year", 0),
                row.get("budget", 0),
                row.get("this_year", 0),
                "AI轉檔"
            ])

    def _process_balance_sheet(self, input_wb, output_wb):
        ws_in = self._find_sheet_by_keyword(input_wb, ["資產", "負債", "平衡"])
        ws_out = output_wb.create_sheet("資產負債表")
        ws_out.append(["科目名稱", "上年度金額", "本年度金額", "說明"])
        
        if not ws_in:
            return

        data = self._extract_data(ws_in, {
            "item": ["科目", "項目", "摘要", "名稱"],
            "this_year": ["本年", "本期", "期末", "決算"],
            "last_year": ["上年", "上期", "期初", "去年"]
        })

        for row in data:
            ws_out.append([
                row.get("item", ""),
                row.get("last_year", 0),
                row.get("this_year", 0),
                "AI轉檔"
            ])
            
    def _process_property_catalog(self, input_wb, output_wb):
         # Create empty or try to find
         ws_out = output_wb.create_sheet("財產目錄")
         ws_out.append(["資產名稱", "取得日期", "取得成本", "本期折舊", "帳面價值"])
         # Simple copy logic omitted for brevity, just ensuring sheet exists
