import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_full_template(filename="Full_Financial_Report_Template.xlsx"):
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    def setup_sheet(sheet_name, headers, sample_data=None):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            
        # Write sample data if provided
        if sample_data:
            for row_num, row_data in enumerate(sample_data, 2):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=cell_value)
                    
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 20
            
        return ws

    # 1. Income Statement (收支決算表)
    income_headers = ["科目名稱", "上年度決算數", "本年度預算數", "本年度決算數", "說明"]
    income_data = [
        ["收入總額", 1000000, 1100000, 1050000, "系統自動計算"],
        ["  會費收入", 200000, 220000, 210000, ""],
        ["  捐款收入", 500000, 500000, 400000, ""],
        ["  補助收入", 300000, 380000, 440000, ""],
        ["支出總額", 800000, 900000, 850000, "系統自動計算"],
        ["  人事費", 400000, 450000, 420000, ""],
        ["  辦公費", 100000, 120000, 110000, ""],
        ["  業務費", 290000, 310000, 290000, ""],
        ["  折舊費用", 10000, 20000, 30000, "固定資產提列"], # Matches "Depreciation" check
        ["本期餘絀", 200000, 200000, 200000, "系統自動計算 (收入-支出)"]
    ]
    ws_income = setup_sheet("收支決算表", income_headers, income_data)
    # Remove default sheet if exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # 2. Balance Sheet (資產負債表)
    balance_headers = ["科目名稱", "上年度金額", "本年度金額", "說明"]
    balance_data = [
        ["資產總額", 2000000, 2200000, "系統自動計算"],
        ["  流動資產", 1500000, 1700000, ""],
        ["    現金及銀行存款", 1400000, 1600000, ""],
        ["    應收帳款", 100000, 100000, ""],
        ["  固定資產", 500000, 500000, ""],
        ["    土地", 400000, 400000, ""],
        ["    房屋", 100000, 100000, ""],
        ["負債總額", 500000, 500000, "系統自動計算"],
        ["  流動負債", 300000, 300000, ""],
        ["    應付帳款", 300000, 300000, ""],
        ["  長期負債", 200000, 200000, ""],
        ["基金及餘絀總額", 1500000, 1700000, "系統自動計算 (資產-負債)"],
        ["  基金", 1000000, 1000000, ""],
        ["  累積餘絀", 300000, 500000, "上期累積餘絀 + 本期餘絀"],
        ["  本期餘絀", 200000, 200000, "應與收支決算表一致"]
    ]
    setup_sheet("資產負債表", balance_headers, balance_data)

    # 3. Fund Income Statement (基金收支表)
    fund_headers = ["科目名稱", "上年度金額", "本年度金額", "說明"]
    fund_data = [
        ["基金收入", 0, 0, ""],
        ["基金支出", 0, 0, ""],
        ["本期基金餘絀", 0, 0, ""]
    ]
    setup_sheet("基金收支表", fund_headers, fund_data)
    
    # 4. Property Catalog (財產目錄) - for Fixed Assets check
    property_headers = ["資產名稱", "取得日期", "取得成本", "本期折舊", "帳面價值"]
    property_data = [
        ["土地", "2010-01-01", 400000, 0, 400000],
        ["房屋", "2015-01-01", 200000, 10000, 100000] # Assuming accumulated depreciation makes it 100k
    ]
    setup_sheet("財產目錄", property_headers, property_data)

    wb.save(filename)
    print(f"Generated template: {filename}")

if __name__ == "__main__":
    create_full_template()
