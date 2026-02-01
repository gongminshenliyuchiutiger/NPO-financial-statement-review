import openpyxl
import sys

def verify_report(filename="Full_Financial_Report_Template.xlsx", output_file="verification_report.txt"):
    
    # Custom print function to write to both console and file
    def log(message):
        print(message)
        with open(output_file, "a", encoding="utf-8") as f:
            f.write(message + "\n")

    # Clear previous report
    with open(output_file, "w", encoding="utf-8") as f:
        f.write("")

    log(f"開始檢核財務報表: {filename}")
    log("-" * 50)
    
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
    except Exception as e:
        log(f"無法讀取檔案: {e}")
        return

    # Helper function to find value by label in a sheet
    def get_value_by_label(sheet_name, label_col_idx, value_col_idx, target_labels):
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[label_col_idx] and any(label in str(row[label_col_idx]) for label in target_labels):
                return row[value_col_idx]
        return 0

    # Helper to get all row data
    def get_rows(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return list(ws.iter_rows(min_row=2, values_only=True))

    # --- 1. 收支決算表檢核 ---
    log("\n【收支決算表檢核】")
    income_rows = get_rows("收支決算表")
    if not income_rows:
        log("[ERROR] 找不到 '收支決算表' 工作表")
    else:
        # P.S. Adjust indices based on generator: 
        # Col 0: Name, Col 1: Last Year, Col 2: Budget, Col 3: This Year Final
        
        income_total = 0
        expense_total = 0
        surplus = 0
        depreciation = 0
        
        # Parse data
        for row in income_rows:
            name = str(row[0]).strip() if row[0] else ""
            val_final = row[3] if isinstance(row[3], (int, float)) else 0
            val_budget = row[2] if isinstance(row[2], (int, float)) else 0
            val_last = row[1] if isinstance(row[1], (int, float)) else 0
            
            if "收入總額" in name: income_total = val_final
            if "支出總額" in name: expense_total = val_final
            if "本期餘絀" in name: surplus = val_final
            if "折舊" in name: depreciation += val_final

            # Rule: Budget vs Final Diff > 20%
            if val_budget > 0 and abs(val_final - val_budget) / val_budget > 0.2:
                log(f"[!] [預算差異] {name}: 決算({val_final}) 與預算({val_budget}) 差異超過 20%")

            # Rule: Year-over-Year Diff > 10%
            if val_last > 0 and abs(val_final - val_last) / val_last > 0.1:
                log(f"[!] [年度變動] {name}: 本年({val_final}) 與上年({val_last}) 變動超過 10%")
            
            # Rule: Single Expense > 20% of Total Expense (Exclude "支出總額" itself)
            if "支出總額" not in name and "收入" not in name and "餘絀" not in name and expense_total > 0:
                 pass 

        # 1. Balance Check
        if abs(income_total - (expense_total + surplus)) > 1:
            log(f"[ERROR] [報表平衡] 收支不平衡: 收入({income_total}) != 支出({expense_total}) + 餘絀({surplus})")
        else:
            log(f"[OK] [報表平衡] 收支平衡正確")

        # 2. Surplus Ratio Check
        if income_total > 0 and (surplus / income_total) > 0.4:
            log(f"[!] [結餘過高] 本期餘絀佔收入 {surplus/income_total:.1%} (標準: <40%) -> 請說明結餘用途")
        else:
            log(f"[OK] [結餘比率] 正常")

    # --- 2.資產負債表檢核 ---
    log("\n【資產負債表檢核】")
    balance_rows = get_rows("資產負債表")
    if not balance_rows:
        log("[ERROR] 找不到 '資產負債表' 工作表")
    else:
        # Col 0: Name, Col 2: This Year
        asset_total = 0
        liability_total = 0
        fund_surplus_total = 0
        bs_surplus = 0
        fixed_assets = 0
        
        for row in balance_rows:
            name = str(row[0]).strip() if row[0] else ""
            val = row[2] if isinstance(row[2], (int, float)) else 0
            
            if "資產總額" in name: asset_total = val
            if "負債總額" in name: liability_total = val
            if "基金及餘絀總額" in name: fund_surplus_total = val
            if "本期餘絀" in name: bs_surplus = val
            if "固定資產" in name and "總額" not in name: fixed_assets = val 

        # 1. Balance Check
        if abs(asset_total - (liability_total + fund_surplus_total)) > 1:
             log(f"[ERROR] [報表平衡] 資產負債不平衡: 資產({asset_total}) != 負債({liability_total}) + 基金餘絀({fund_surplus_total})")
        else:
             log(f"[OK] [報表平衡] 資產負債平衡正確")

        # 2. Cross-Sheet Consistnecy (Surplus)
        if abs(bs_surplus - surplus) > 1:
             log(f"[ERROR] [跨表勾稽] 資產負債表餘絀({bs_surplus}) 與 收支決算表餘絀({surplus}) 不符")
        else:
             log(f"[OK] [跨表勾稽] 餘絀金額一致")
             
        # 3. Debt Ratio
        if asset_total > 0 and (liability_total / asset_total) > 0.5:
            log(f"[!] [財務結構] 負債比率 {liability_total/asset_total:.1%} (標準: <50%) -> 請關注償債能力")
        else:
            log(f"[OK] [財務結構] 負債比率正常")

        # 4. Depreciation Check Logic
        if fixed_assets > 0 and depreciation == 0:
             log(f"[!] [折舊查核]帳上有固定資產({fixed_assets})，但收支表無折舊費用")

    # --- 3. 財產目錄勾稽 ---
    log("\n【財產目錄勾稽】")
    prop_rows = get_rows("財產目錄")
    if prop_rows:
        # Col 4: Book Value
        prop_total_value = sum(row[4] for row in prop_rows if isinstance(row[4], (int, float)))
        if abs(fixed_assets - prop_total_value) > 1:
             log(f"[ERROR] [財產目錄] 固定資產帳面價值({fixed_assets}) 與 財產目錄加總({prop_total_value}) 不符")
        else:
             log(f"[OK] [財產目錄] 金額相符")
    else:
        log("[!] 無財產目錄資料")

    log("\n檢核完成")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        verify_report(sys.argv[1])
    else:
        verify_report()
