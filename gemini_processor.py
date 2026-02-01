import google.generativeai as genai
import openpyxl
import json
import os
import time

class GeminiProcessor:
    def __init__(self, api_key):
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-1.5-flash')

    def process_file(self, filepath, mime_type):
        # 1. Upload file to Gemini
        print(f"Uploading file {filepath} to Gemini...")
        uploaded_file = genai.upload_file(path=filepath, mime_type=mime_type)
        
        # Wait for processing if needed (mostly for videos, but safe for large PDFs)
        while uploaded_file.state.name == "PROCESSING":
            time.sleep(1)
            uploaded_file = genai.get_file(uploaded_file.name)

        if uploaded_file.state.name == "FAILED":
            raise ValueError("Gemini failed to process the file.")

        # 2. Construct Prompt
        prompt = """
        你是專業的非營利組織財務會計師。請協助將上傳的財務報表文件（可能是圖片、PDF或Excel），轉換為標準的 JSON 格式。
        
        請提取以下三張表的數據：
        1. 收支決算表 (Income Statement): 包含科目名稱(item)、上年度決算數(last_year)、本年度預算數(budget)、本年度決算數(this_year)
        2. 資產負債表 (Balance Sheet): 包含科目名稱(item)、上年度金額(last_year)、本年度金額(this_year)
        3. 財產目錄 (Property Catalog): 包含資產名稱(item_name)、取得日期(date)、取得成本(cost)、本期折舊(depreciation)、帳面價值(book_value)

        請務必返回純 JSON 格式，不要有 Markdown 標記（如 ```json ... ```）。結構如下：
        {
            "income_statement": [
                {"item": "收入總額", "last_year": 1000, "budget": 1200, "this_year": 1100},
                ...
            ],
            "balance_sheet": [
                {"item": "資產總額", "last_year": 5000, "this_year": 5500},
                ...
            ],
            "property_catalog": [
                {"item_name": "電腦", "date": "2023-01-01", "cost": 30000, "depreciation": 10000, "book_value": 20000},
                ...
            ]
        }
        
        如果文件中缺某些欄位（例如沒有預算數），請填 0 或 null。請自動識別並對應到最接近的標準科目名稱。
        """

        # 3. Generate Content
        print("Analyzing with Gemini...")
        response = self.model.generate_content([prompt, uploaded_file])
        
        # Cleanup file from cloud
        uploaded_file.delete()
        
        # 4. Parse JSON
        try:
            # Strip markdown if present
            text = response.text.strip()
            if text.startswith("```json"):
                text = text[7:]
            if text.endswith("```"):
                text = text[:-3]
            data = json.loads(text)
            return self._json_to_workbook(data)
        except Exception as e:
            print(f"Error parsing Gemini response: {e}")
            print(f"Raw response: {response.text}")
            raise

    def _json_to_workbook(self, data):
        wb = openpyxl.Workbook()
        
        # 1. Income Statement
        ws_income = wb.create_sheet("收支決算表")
        ws_income.append(["科目名稱", "上年度決算數", "本年度預算數", "本年度決算數", "說明"])
        if "income_statement" in data:
            for row in data["income_statement"]:
                ws_income.append([
                    row.get("item", ""),
                    row.get("last_year", 0),
                    row.get("budget", 0),
                    row.get("this_year", 0),
                    "Gemini擷取"
                ])

        # 2. Balance Sheet
        ws_balance = wb.create_sheet("資產負債表")
        ws_balance.append(["科目名稱", "上年度金額", "本年度金額", "說明"])
        if "balance_sheet" in data:
            for row in data["balance_sheet"]:
                ws_balance.append([
                    row.get("item", ""),
                    row.get("last_year", 0),
                    row.get("this_year", 0),
                    "Gemini擷取"
                ])

        # 3. Property Catalog
        ws_prop = wb.create_sheet("財產目錄")
        ws_prop.append(["資產名稱", "取得日期", "取得成本", "本期折舊", "帳面價值"])
        if "property_catalog" in data:
             for row in data["property_catalog"]:
                ws_prop.append([
                    row.get("item_name", ""),
                    row.get("date", ""),
                    row.get("cost", 0),
                    row.get("depreciation", 0),
                    row.get("book_value", 0)
                ])

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
            
        return wb
