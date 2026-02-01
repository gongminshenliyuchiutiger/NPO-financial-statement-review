import openpyxl
import sys

# Force utf-8 for stdout if needed, but writing to file is safer to avoid console encoding issues
output_file = 'excel_structure.txt'

try:
    wb = openpyxl.load_workbook('AI_財報自動檢核範本.xlsx', data_only=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(f"Workbook: AI_財報自動檢核範本.xlsx\n")
        f.write("-" * 20 + "\n")
        
        for sheet_name in wb.sheetnames:
            f.write(f"\nSheet: {sheet_name}\n")
            ws = wb[sheet_name]
            f.write(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
            
            # Print first 10 rows to understand headers and data location
            f.write("First 10 rows content:\n")
            for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
                # Filter out None values for cleaner output
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                f.write(f"Row {i}: {cleaned_row}\n")

    print(f"Structure saved to {output_file}")

except Exception as e:
    print(f"Error: {e}")
