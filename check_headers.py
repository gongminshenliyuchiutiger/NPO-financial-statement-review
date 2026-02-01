import openpyxl
import sys

# Force utf-8 encoding for stdout
sys.stdout.reconfigure(encoding='utf-8')

try:
    wb = openpyxl.load_workbook('NPO_財務報表檢核範本_v2.xlsx')
    for sheet in wb.sheetnames:
        print(f"=== Sheet: {repr(sheet)} ===")
        ws = wb[sheet]
        row1 = [repr(cell.value) for cell in ws[1]]
        print(f"Headers: {row1}")
        # Also check row 2 for any strict formulas or data types if needed
        # row2 = [repr(cell.value) for cell in ws[2]]
        # print(f"Row 2: {row2}")
except Exception as e:
    print(f"Error: {e}")
