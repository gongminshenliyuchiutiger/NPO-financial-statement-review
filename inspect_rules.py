import json
import openpyxl
import sys

# Force utf-8 encoding for stdout
sys.stdout.reconfigure(encoding='utf-8')

def extract_rules_to_json(filename):
    rules = []
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        # Look for sheet that likely contains the rules
        target_sheet = None
        for s in wb.sheetnames:
            if '說明' in s or '檢核' in s:
                target_sheet = s
                break
        
        if not target_sheet:
            print(f"No rule sheet found in {filename}")
            return

        ws = wb[target_sheet]
        rows = list(ws.iter_rows(values_only=True))
        
        # Assume Header is Row 1 (Index 0)
        # Headers: Item, Content, Meaning, Risk, Criteria, Suggestion
        # Let's map dynamically if possible, or assume fixed index based on previous view
        # Previous view: Item(0), Content(1), Meaning(2), Risk(3), Normal(4), Suggestion(5)
        
        for i, row in enumerate(rows):
            if i == 0: continue # Skip header
            if not row[0]: continue # Skip empty rows
            
            rule = {
                "Item": str(row[0]).strip(),
                "Content": str(row[1]).strip() if len(row)>1 else "",
                "Meaning": str(row[2]).strip() if len(row)>2 else "",    # Positive Meaning
                "Risk": str(row[3]).strip() if len(row)>3 else "",       # Abnormal Meaning
                "Criteria": str(row[4]).strip() if len(row)>4 else "",
                "Suggestion": str(row[5]).strip() if len(row)>5 else ""
            }
            rules.append(rule)
            
        print(f"=== Rules from {filename} ===")
        print(json.dumps(rules, ensure_ascii=False, indent=2))

    except Exception as e:
        print(f"Error reading {filename}: {e}")

files = [
    '財務報表_收支決算表檢核說明.xlsx',
    '財務報表_資產負債表檢核說明.xlsx'
]

for f in files:
    extract_rules_to_json(f)
