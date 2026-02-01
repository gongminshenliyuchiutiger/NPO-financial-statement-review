import openpyxl
import sys

# Force utf-8 for stdout
output_file = 'excel_structure_v2.txt'

try:
    # Changed file name to v2
    wb = openpyxl.load_workbook('NPO_財務報表檢核範本_v2.xlsx', data_only=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(f"Workbook: NPO_財務報表檢核範本_v2.xlsx\n")
        f.write("-" * 20 + "\n")
        
        for sheet_name in wb.sheetnames:
            f.write(f"\nSheet: {sheet_name}\n")
            ws = wb[sheet_name]
            f.write(f"Max Row: {ws.max_row}, Max Column: {ws.max_column}\n")
            
            # Print first 15 rows to cover more headers
            f.write("First 15 rows content:\n")
            for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                f.write(f"Row {i}: {cleaned_row}\n")

    print(f"Structure saved to {output_file}")

except Exception as e:
    print(f"Error: {e}")
